"""
skills/silac_analysis.py
SILAC (Stable Isotope Labelling by Amino acids in Cell culture) analysis skill.

SILAC data characteristics
--------------------------
• Values are H/L, H/M, or M/L ratios (or their log2 equivalents).
• No sample-level replicates for normalisation — biological replicates are
  separate experiments with their own H/L columns.
• Normalisation: median ratio per sample → subtract from all ratios (log2 space).
• Single-condition SILAC: one-sample t-test vs. 0 (i.e. ratio ≠ 1).
• Multi-condition SILAC: compare log2 ratios between two experimental sets
  using a standard Welch t-test on the ratio values.

Detection heuristics
---------------------
Spectral data with column names containing "ratio", "h/l", "h/m", "m/l",
"heavy", "light", or "medium" is treated as SILAC.
Values in [-7, 7] that are not all integers are assumed to be pre-log2 ratios;
values outside that range are raw ratios and are log2-transformed.
"""
from __future__ import annotations

import logging
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from scipy import stats
from statsmodels.stats.multitest import multipletests

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from skills.base_skill import BaseOmicsSkill, OmicsAnalysisResult

logger = logging.getLogger(__name__)

_SILAC_RATIO_KEYWORDS = (
    "ratio h/l", "ratio h/m", "ratio m/l",
    "ratio hl", "ratio hm", "ratio ml",
    "h/l ratio", "h/m ratio", "m/l ratio",
    "heavy/light", "heavy/medium", "medium/light",
    "silac ratio", "normalized ratio",
)
_SILAC_COL_HINTS = ("ratio", "h/l", "h/m", "m/l", "heavy", "light", "medium", "silac")


def is_silac_data(df: pd.DataFrame, sample_cols: List[str]) -> bool:
    """Return True when column names or index suggest SILAC ratio data."""
    col_names = " ".join(str(c).lower() for c in df.columns)
    idx_names  = " ".join(str(i).lower() for i in df.index[:20])
    combined   = col_names + " " + idx_names
    return any(kw in combined for kw in _SILAC_COL_HINTS)


class SilacAnalysisSkill(BaseOmicsSkill):
    """
    SILAC differential analysis.
    Handles:
      • Single-condition SILAC — one-sample t-test vs 0 (are log2 ratios ≠ 0?)
      • Two-condition SILAC    — Welch t-test comparing two sets of log2 ratios
    Produces: BH-corrected p-values, log2 ratio means, Cohen's d, Excel report.
    """

    @property
    def omic_type(self) -> str:
        return "proteomics_silac"

    def execute(  # type: ignore[override]
        self,
        data_path: str,
        sample_columns: List[str],
        group1_samples: Optional[List[str]] = None,
        group2_samples: Optional[List[str]] = None,
        group1_label: str = "Condition1",
        group2_label: str = "Condition2",
        analysis_mode: str = "supervised",
        data_type: str = "silac_ratio",
        adj_pval_cutoff: float = 0.05,
        log2fc_cutoff: float = 1.0,
        missing_threshold: float = 0.5,
        top_n: int = 50,
        output_dir: str = "outputs",
        file_name: str = "analysis",
        **_kwargs,
    ) -> Dict[str, Any]:
        try:
            result = self._run(
                data_path, sample_columns, group1_samples, group2_samples,
                group1_label, group2_label, analysis_mode,
                adj_pval_cutoff, log2fc_cutoff, missing_threshold,
                top_n, output_dir, file_name,
            )
            result["omic_type"] = self.omic_type
            return result
        except Exception as exc:
            logger.error("SilacAnalysisSkill failed:\n%s", traceback.format_exc())
            return OmicsAnalysisResult(
                omic_type=self.omic_type,
                top_biomarkers=[],
                n_significant=0,
                excel_path=None,
                qc_summary={},
                error=str(exc),
            )

    # ── Internal pipeline ──────────────────────────────────────────────────────

    def _run(
        self,
        data_path, sample_columns, group1_samples, group2_samples,
        group1_label, group2_label, analysis_mode,
        adj_pval_cutoff, log2fc_cutoff, missing_threshold,
        top_n, output_dir, file_name,
    ) -> Dict[str, Any]:

        df_raw = pd.read_csv(data_path, index_col=0)
        avail  = [c for c in sample_columns if c in df_raw.columns]
        if not avail:
            raise ValueError(
                "None of the specified sample_columns found in the SILAC data file."
            )
        data = df_raw[avail].apply(pd.to_numeric, errors="coerce")

        data, qc_summary = self._qc_silac(data, missing_threshold)

        if (
            analysis_mode == "supervised"
            and group1_samples
            and group2_samples
        ):
            results_df = self._two_condition(
                data, group1_samples, group2_samples,
                group1_label, group2_label,
                adj_pval_cutoff, log2fc_cutoff, missing_threshold,
            )
        else:
            results_df = self._one_condition(
                data, adj_pval_cutoff, log2fc_cutoff, missing_threshold,
            )

        sig_mask = (
            (results_df["adj_p_value"] < adj_pval_cutoff)
            & (results_df["log2_ratio"].abs() >= log2fc_cutoff)
        )
        n_sig = int(sig_mask.sum())
        top_df = results_df.head(top_n).copy()

        Path(output_dir).mkdir(parents=True, exist_ok=True)
        ts         = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name  = Path(file_name).stem.replace(" ", "_")
        excel_path = str(Path(output_dir) / f"silac_{safe_name}_{ts}.xlsx")
        self._export_excel(results_df, top_df, qc_summary, excel_path,
                           group1_label, group2_label, analysis_mode,
                           adj_pval_cutoff, log2fc_cutoff, file_name)

        return {
            "omic_type":      self.omic_type,
            "top_biomarkers": top_df.to_dict("records"),
            "n_significant":  n_sig,
            "excel_path":     excel_path,
            "qc_summary":     qc_summary,
            "analysis_code":  None,
            "error":          None,
        }

    # ── QC for SILAC ratios ────────────────────────────────────────────────────

    @staticmethod
    def _qc_silac(
        data: pd.DataFrame,
        missing_threshold: float,
    ) -> Tuple[pd.DataFrame, Dict]:
        orig_proteins = len(data)
        orig_samples  = len(data.columns)

        # Detect whether values are raw ratios (> 0, often > 1) or log2 ratios
        col_max = data.max().max()
        col_min = data.min().min()
        is_raw  = (
            pd.notna(col_max)
            and pd.notna(col_min)
            and float(col_min) > 0
            and float(col_max) > 10
        )
        if is_raw:
            data = np.log2(data.replace(0, np.nan))
            log2_done = True
        else:
            log2_done = False

        # Global protein filter
        data = data.loc[data.isna().mean(axis=1) <= missing_threshold]

        # Median normalisation in log2 space (subtract per-sample median)
        sample_medians = data.median(axis=0, skipna=True)
        data = data.subtract(sample_medians, axis=1)

        # Impute missing with 0 (log2 ratio = 0 means no change — conservative)
        data = data.fillna(0.0)

        return data, {
            "proteins_input":       orig_proteins,
            "proteins_after_qc":    len(data),
            "samples_input":        orig_samples,
            "log2_transformed":     log2_done,
            "normalisation_method": "median_ratio",
            "imputation_method":    "zero (no change assumed)",
            "missing_threshold":    missing_threshold,
        }

    # ── One-condition SILAC: one-sample t-test vs 0 ────────────────────────────

    @staticmethod
    def _one_condition(
        data: pd.DataFrame,
        adj_pval_cutoff: float,
        log2fc_cutoff: float,
        missing_threshold: float,
    ) -> pd.DataFrame:
        rows = []
        for protein in data.index:
            v = data.loc[protein].values.astype(float)
            v = v[v != 0.0]                   # exclude imputed zeros
            if len(v) < 2:
                continue
            _, pval = stats.ttest_1samp(v, 0.0)
            if np.isnan(pval):
                continue
            mean_ratio = float(v.mean())
            sd         = float(v.std(ddof=1))
            cohens_dz  = mean_ratio / sd if sd > 0 else 0.0
            rows.append({
                "protein":    protein,
                "log2_ratio": round(mean_ratio, 4),
                "cohens_dz":  round(cohens_dz, 4),
                "p_value":    float(pval),
                "n_replicates": len(v),
                "sd_ratio":   round(sd, 4),
            })

        if not rows:
            raise ValueError("No proteins had ≥2 non-zero SILAC ratios for testing.")
        df_res = pd.DataFrame(rows)
        _, adj_p, _, _ = multipletests(df_res["p_value"].values, method="fdr_bh")
        df_res["adj_p_value"] = adj_p

        hi_pval    = min(0.01, adj_pval_cutoff / 5.0)
        trend_pval = adj_pval_cutoff * 2.0
        df_res["significance"] = "NS"
        hi  = (df_res["adj_p_value"] < hi_pval)        & (df_res["log2_ratio"].abs() >= log2fc_cutoff)
        sig = (df_res["adj_p_value"] < adj_pval_cutoff) & (df_res["log2_ratio"].abs() >= log2fc_cutoff)
        trn = (df_res["adj_p_value"] < trend_pval)      & ~sig
        df_res.loc[trn, "significance"] = "Trend"
        df_res.loc[sig, "significance"] = "Significant"
        df_res.loc[hi,  "significance"] = "Highly Significant"
        df_res = df_res.sort_values("adj_p_value").reset_index(drop=True)
        df_res.insert(0, "rank", range(1, len(df_res) + 1))
        return df_res

    # ── Two-condition SILAC: Welch t-test on ratio sets ─────────────────────────

    @staticmethod
    def _two_condition(
        data: pd.DataFrame,
        g1_cols: List[str],
        g2_cols: List[str],
        g1_label: str,
        g2_label: str,
        adj_pval_cutoff: float,
        log2fc_cutoff: float,
        missing_threshold: float,
    ) -> pd.DataFrame:
        g1 = [c for c in g1_cols if c in data.columns]
        g2 = [c for c in g2_cols if c in data.columns]
        if len(g1) < 2 or len(g2) < 2:
            raise ValueError(
                f"Two-condition SILAC requires ≥2 columns per group. "
                f"Got: {g1_label}={len(g1)}, {g2_label}={len(g2)}."
            )

        rows = []
        for protein in data.index:
            v1 = data.loc[protein, g1].values.astype(float)
            v2 = data.loc[protein, g2].values.astype(float)
            if len(v1) < 2 or len(v2) < 2:
                continue
            m1, m2 = v1.mean(), v2.mean()
            _, pval = stats.ttest_ind(v1, v2, equal_var=False)
            if np.isnan(pval):
                continue
            n1, n2 = len(v1), len(v2)
            s1, s2 = v1.std(ddof=1), v2.std(ddof=1)
            sp = np.sqrt(((n1 - 1) * s1**2 + (n2 - 1) * s2**2) / max(n1 + n2 - 2, 1))
            rows.append({
                "protein":              protein,
                f"mean_{g1_label}":     round(float(m1), 4),
                f"mean_{g2_label}":     round(float(m2), 4),
                "log2_ratio":           round(float(m1 - m2), 4),
                "cohens_d":             round(float((m1 - m2) / sp) if sp > 0 else 0.0, 4),
                "p_value":              float(pval),
                f"n_{g1_label}":        n1,
                f"n_{g2_label}":        n2,
            })

        if not rows:
            raise ValueError("No proteins had sufficient data for two-condition SILAC analysis.")
        df_res = pd.DataFrame(rows)
        _, adj_p, _, _ = multipletests(df_res["p_value"].values, method="fdr_bh")
        df_res["adj_p_value"] = adj_p

        hi_pval    = min(0.01, adj_pval_cutoff / 5.0)
        trend_pval = adj_pval_cutoff * 2.0
        df_res["significance"] = "NS"
        hi  = (df_res["adj_p_value"] < hi_pval)        & (df_res["log2_ratio"].abs() >= log2fc_cutoff)
        sig = (df_res["adj_p_value"] < adj_pval_cutoff) & (df_res["log2_ratio"].abs() >= log2fc_cutoff)
        trn = (df_res["adj_p_value"] < trend_pval)      & ~sig
        df_res.loc[trn, "significance"] = "Trend"
        df_res.loc[sig, "significance"] = "Significant"
        df_res.loc[hi,  "significance"] = "Highly Significant"
        df_res = df_res.sort_values("adj_p_value").reset_index(drop=True)
        df_res.insert(0, "rank", range(1, len(df_res) + 1))
        return df_res

    # ── Excel export ───────────────────────────────────────────────────────────

    def _export_excel(
        self,
        all_results: pd.DataFrame,
        top_results: pd.DataFrame,
        qc_summary: Dict,
        output_path: str,
        g1_label: str,
        g2_label: str,
        analysis_mode: str,
        adj_pval_cutoff: float,
        log2fc_cutoff: float,
        file_name: str,
    ) -> None:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            top_results.to_excel(writer, sheet_name="Top Proteins",  index=False)
            all_results.to_excel(writer, sheet_name="All Results",   index=False)
            pd.DataFrame([{"Parameter": k.replace("_", " ").title(), "Value": v}
                          for k, v in qc_summary.items()]).to_excel(
                writer, sheet_name="QC Summary", index=False)
            params = {
                "Analysis Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "File":          file_name,
                "Mode":          "SILAC " + ("two-condition" if analysis_mode == "supervised"
                                              else "one-condition"),
                "Test":          ("Welch t-test on log₂ ratios" if analysis_mode == "supervised"
                                   else "One-sample t-test vs 0"),
                "FDR Method":    "Benjamini-Hochberg",
                "Adj P Cutoff":  adj_pval_cutoff,
                "Log₂ Ratio Cutoff": log2fc_cutoff,
            }
            pd.DataFrame([{"Parameter": k, "Value": v}
                          for k, v in params.items()]).to_excel(
                writer, sheet_name="Parameters", index=False)

        # Basic header formatting
        wb = openpyxl.load_workbook(output_path)
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                    max((len(str(c.value)) for c in col if c.value is not None), default=8) + 4, 40
                )
            ws.freeze_panes = "A2"
        wb.save(output_path)
