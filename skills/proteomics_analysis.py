"""
skills/proteomics_analysis.py
Analysis Layer — ProteomicsAnalysisSkill (intensity-only, canonical template)

Pipeline
--------
1. QC          : log2 transform → contaminant removal → global protein/sample filter
2. Normalise   : median-centering (removes between-sample intensity shifts)
               : IRS normalisation for multi-batch TMT (when tmt_batches provided)
3. Filter      : group-aware missing-value filter (protein must be detected in
                 ≥50 % of samples in at least one comparison group)
4. Impute      : half-minimum per protein (Perseus-style)
5. Stats       : auto-selected or user-requested test method —
                   "welch"    — Welch two-sample t-test          (n ≥ 5)
                   "limma"    — empirical Bayes moderated t-test  (n ≤ 4, recommended)
                   "paired_t" — paired t-test (matched samples, before/after)
                   "anova"    — one-way ANOVA for >2 groups
                 or CV / MAD ranking (unsupervised, no group labels)
6. FDR         : Benjamini-Hochberg correction
7. Effect size : log₂FC + Cohen's d (or Cohen's d_z for paired, η² for ANOVA)
8. Export      : formatted multi-sheet Excel workbook

This is the Python half of the dual-engine pipeline. The R half lives in
``skills/r_analysis.py``; results are intersected by
``skills/dual_engine.py`` to produce the final top-biomarker list.
"""
from __future__ import annotations

import logging
import traceback
from datetime import datetime

logger = logging.getLogger(__name__)
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from scipy import stats
from scipy.special import digamma, polygamma
from scipy.optimize import brentq
from statsmodels.stats.multitest import multipletests

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from skills.base_skill import BaseOmicsSkill, OmicsAnalysisResult


# ── Colour palette ────────────────────────────────────────────────────────────
_HEADER_FG  = "FFFFFF"
_HEADER_BG  = "1F4E79"
_HIGHLY_SIG = "00B050"   # dark green
_SIGNIFICANT = "92D050"  # light green
_TREND       = "FFEB9C"  # pale yellow
_VARIABLE    = "BDD7EE"  # pale blue (unsupervised top proteins)


def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style() -> Tuple[PatternFill, Font]:
    fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    font = Font(bold=True, color=_HEADER_FG, size=11)
    return fill, font


class ProteomicsAnalysisSkill(BaseOmicsSkill):
    """
    General-purpose proteomics biomarker discovery skill.

    Handles both replicated designs (≥2 samples per group) via Welch t-test
    and exploratory (no group labels) via CV ranking.  All analysis is driven
    purely by the data uploaded in the session — no hardcoded group names,
    sample names, or thresholds.
    """

    @property
    def omic_type(self) -> str:
        return "proteomics"

    def execute(  # type: ignore[override]
        self,
        data_path: str,
        sample_columns: List[str],
        group1_samples: Optional[List[str]] = None,
        group2_samples: Optional[List[str]] = None,
        group1_label: str = "Group1",
        group2_label: str = "Group2",
        analysis_mode: str = "supervised",
        data_type: str = "generic",
        adj_pval_cutoff: float = 0.05,
        log2fc_cutoff: float = 1.0,
        missing_threshold: float = 0.5,
        top_n: int = 50,
        output_dir: str = "outputs",
        file_name: str = "analysis",
        # Extended test method selection
        test_method: str = "auto",          # "auto"|"welch"|"limma"|"paired_t"|"anova"|"fold_change"
        is_paired: bool = False,
        all_groups: Optional[Dict[str, List[str]]] = None,
        tmt_batches: Optional[Dict[str, Any]] = None,
        is_pooled_design: bool = False,
        **_kwargs,
    ) -> Dict[str, Any]:
        try:
            result = self._run(
                data_path, sample_columns, group1_samples, group2_samples,
                group1_label, group2_label, analysis_mode, data_type,
                adj_pval_cutoff, log2fc_cutoff, missing_threshold,
                top_n, output_dir, file_name,
                test_method=test_method,
                is_paired=is_paired,
                all_groups=all_groups,
                tmt_batches=tmt_batches,
                is_pooled_design=is_pooled_design,
            )
            result["omic_type"] = self.omic_type
            return result
        except Exception as exc:
            logger.error("ProteomicsAnalysisSkill failed:\n%s", traceback.format_exc())
            return OmicsAnalysisResult(
                omic_type=self.omic_type,
                top_biomarkers=[],
                n_significant=0,
                excel_path=None,
                qc_summary={},
                error=str(exc),
            )

    # ── Internal pipeline ─────────────────────────────────────────────────────

    def _run(
        self,
        data_path, sample_columns, group1_samples, group2_samples,
        group1_label, group2_label, analysis_mode, data_type,
        adj_pval_cutoff, log2fc_cutoff, missing_threshold,
        top_n, output_dir, file_name,
        test_method: str = "auto",
        is_paired: bool = False,
        all_groups: Optional[Dict[str, List[str]]] = None,
        tmt_batches: Optional[Dict[str, Any]] = None,
        is_pooled_design: bool = False,
    ) -> Dict[str, Any]:

        # 1. Load
        df_raw = pd.read_csv(data_path, index_col=0)
        avail = [c for c in sample_columns if c in df_raw.columns]
        if not avail:
            raise ValueError(
                "None of the specified sample_columns are present in the data file.\n"
                f"Available columns: {list(df_raw.columns[:10])}"
            )
        data = df_raw[avail].apply(pd.to_numeric, errors="coerce")

        # Protein-metadata lookup — the loaded CSV carries the protein-name
        # and gene-name columns alongside the numeric samples. We capture them
        # here so every result row can be enriched with the *actual* names
        # from the user's file. Without this, downstream LLM consumers (the
        # domain-expert prompt, the summary message) only see accession
        # numbers and end up hallucinating protein identities.
        meta_lookup = self._build_protein_metadata(df_raw, avail)

        # 2. QC + normalisation + optional TMT IRS
        required_samples = list(group1_samples or []) + list(group2_samples or [])
        if all_groups:
            for cols in all_groups.values():
                required_samples.extend(cols or [])
        required_samples = [c for c in dict.fromkeys(required_samples) if c in data.columns]
        data_qc, qc_summary, valid_mask = self._qc(
            data,
            missing_threshold,
            data_type,
            required_samples=required_samples,
            tmt_batches=tmt_batches,
        )

        # 3. Resolve effective test method
        effective_method = self._resolve_test_method(
            test_method, is_paired,
            group1_samples or [], group2_samples or [],
            data_qc,
        )

        # 4. Analysis branch
        # Pooled / non-replicated designs (1 sample per condition, or pooled
        # MaxQuant Identifier Info workbooks) cannot support a valid t-test
        # because there are no true biological replicates within each "group" —
        # within-group spread is dominated by between-sample variation. We run
        # fold-change-only ranking instead, with NO p-values.
        if (is_pooled_design or test_method == "fold_change"
            ) and group1_samples and group2_samples:
            results_df = self._fold_change_only(
                data_qc, valid_mask,
                group1_samples, group2_samples,
                group1_label, group2_label,
                log2fc_cutoff=log2fc_cutoff,
                missing_threshold=missing_threshold,
            )
            sig_mask = results_df["log2_fold_change"].abs() >= log2fc_cutoff
            n_sig = int(sig_mask.sum())
            effective_method = "fold_change_only"
            qc_summary["test_method"] = effective_method
            qc_summary["pooled_design_warning"] = (
                "Pooled / non-replicated design detected — t-tests would treat "
                "between-sample variation as biological noise. Reporting "
                "log2 fold-change only, no p-values."
            )
            results_df = self._enrich_with_metadata(results_df, meta_lookup)
            top_df = results_df.head(top_n).copy()
            return self._finalise_pooled(
                results_df, top_df, qc_summary, output_dir, file_name,
                group1_label, group2_label, n_sig,
                log2fc_cutoff, adj_pval_cutoff,
                data_path, sample_columns,
                group1_samples, group2_samples,
                missing_threshold, top_n,
            )

        is_anova = (
            effective_method == "anova"
            or (all_groups and len(all_groups) >= 2)
        )

        if is_anova and all_groups and len(all_groups) >= 2:
            results_df = self._anova_multigroup(
                data_qc, valid_mask, all_groups,
                log2fc_cutoff=log2fc_cutoff,
                adj_pval_cutoff=adj_pval_cutoff,
                missing_threshold=missing_threshold,
            )
            sig_mask = (
                (results_df["adj_p_value"] < adj_pval_cutoff)
                & (results_df["max_log2fc"].abs() >= log2fc_cutoff)
            )
            n_sig = int(sig_mask.sum())
        elif analysis_mode == "supervised" and group1_samples and group2_samples:
            kw = dict(
                log2fc_cutoff=log2fc_cutoff,
                adj_pval_cutoff=adj_pval_cutoff,
                missing_threshold=missing_threshold,
            )
            if effective_method == "limma":
                results_df = self._limma_ebayes(
                    data_qc, valid_mask,
                    group1_samples, group2_samples,
                    group1_label, group2_label, **kw,
                )
            elif effective_method == "paired_t":
                results_df = self._paired_ttest(
                    data_qc, valid_mask,
                    group1_samples, group2_samples,
                    group1_label, group2_label, **kw,
                )
            else:
                results_df = self._supervised(
                    data_qc, valid_mask,
                    group1_samples, group2_samples,
                    group1_label, group2_label, **kw,
                )
            sig_mask = (
                (results_df["adj_p_value"] < adj_pval_cutoff)
                & (results_df["log2_fold_change"].abs() >= log2fc_cutoff)
            )
            n_sig = int(sig_mask.sum())
        else:
            results_df = self._unsupervised(data_qc, valid_mask)
            n_sig = min(top_n, len(results_df))
            effective_method = "unsupervised"

        qc_summary["test_method"] = effective_method

        # Attach the protein-name + gene-name columns from the source file —
        # every downstream consumer (Excel sheets, the chatbot summary, the
        # domain-expert prompt) gets the ground-truth identity for each row.
        results_df = self._enrich_with_metadata(results_df, meta_lookup)

        top_df = results_df.head(top_n).copy()
        top_biomarkers = top_df.to_dict("records")

        # 4. Excel export
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = Path(file_name).stem.replace(" ", "_")
        excel_path = str(Path(output_dir) / f"biomarkers_{safe_name}_{ts}.xlsx")

        self._export_excel(
            results_df, top_df, qc_summary, excel_path,
            group1_label, group2_label, analysis_mode,
            adj_pval_cutoff, log2fc_cutoff, file_name,
            test_method=effective_method,
        )

        code = self._generate_code(
            data_path, sample_columns, group1_samples, group2_samples,
            group1_label, group2_label, analysis_mode, data_type,
            adj_pval_cutoff, log2fc_cutoff, missing_threshold, top_n,
            output_dir, file_name,
        )

        return {
            "omic_type":      self.omic_type,
            "top_biomarkers": top_biomarkers,
            "n_significant":  n_sig,
            "excel_path":     excel_path,
            "qc_summary":     qc_summary,
            "analysis_code":  code,
            "error":          None,
        }

    # ── QC + normalisation ────────────────────────────────────────────────────

    def _qc(
        self,
        data: pd.DataFrame,
        missing_threshold: float,
        data_type: str,
        required_samples: Optional[List[str]] = None,
        tmt_batches: Optional[Dict[str, Any]] = None,
    ) -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
        """
        Returns (normalised_imputed_data, qc_summary, valid_mask).
        valid_mask is a boolean DataFrame (same shape as filtered data, before
        imputation) where True means the value was genuinely measured.
        """
        orig_proteins = len(data)
        orig_samples  = len(data.columns)

        # ── Step 0: remove contaminants / reverse decoys ──────────────────────
        #    MaxQuant marks these with REV__ / CON__ prefixes.
        #    Common contaminants (keratin, albumin) are also removed.
        _CONTAM_PREFIXES = ("REV__", "CON__", "##")
        _CONTAM_KEYWORDS = ("keratin", "krt", "bsa", "bovine serum albumin")
        idx_str  = data.index.astype(str)
        is_clean = ~(
            idx_str.str.startswith(_CONTAM_PREFIXES[0])
            | idx_str.str.startswith(_CONTAM_PREFIXES[1])
            | idx_str.str.startswith(_CONTAM_PREFIXES[2])
            | idx_str.str.lower().str.contains(
                "|".join(_CONTAM_KEYWORDS), na=False, regex=True
            )
        )
        n_contam = int((~is_clean).sum())
        data = data[is_clean]
        if n_contam:
            logger.info("Removed %d contaminant/reverse-decoy entries.", n_contam)

        # ── Step 1: log2 transform if values look like raw intensities ────────
        log_done = False
        col_max = data.max().max()
        if pd.notna(col_max) and float(col_max) > 100:
            data = np.log2(data.replace(0, np.nan) + 1)
            log_done = True

        # ── Step 2: global protein filter (>threshold missing across ALL samples)
        miss_prot = data.isna().mean(axis=1)
        data = data.loc[miss_prot <= missing_threshold]

        # ── Step 3: sample filter (>80% missing) ─────────────────────────────
        miss_samp = data.isna().mean(axis=0)
        must_keep = set(required_samples or [])
        keep_cols = [c for c in data.columns if (miss_samp.get(c, 1.0) <= 0.80) or (c in must_keep)]
        data = data.loc[:, keep_cols]

        # ── Step 4: median-centering normalisation ────────────────────────────
        #    Subtract each sample's median and add the global median back.
        #    This corrects systematic intensity shifts between samples without
        #    changing fold-changes, and is the standard in LFQ proteomics.
        norm_done = False
        if not data.empty and len(data.columns) > 1:
            sample_medians = data.median(axis=0, skipna=True)
            global_median  = float(sample_medians.median())
            if pd.notna(global_median):
                data      = data.subtract(sample_medians, axis=1).add(global_median)
                norm_done = True

        # ── Step 4b: IRS normalisation for multi-batch TMT ────────────────────
        irs_done = False
        if tmt_batches and len(tmt_batches) >= 2:
            data     = self._apply_irs(data, tmt_batches)
            irs_done = True

        # ── Step 5: save valid mask BEFORE imputation ─────────────────────────
        #    Captures which values are genuinely measured (vs. to be imputed).
        #    Used downstream to compute per-group detection rates.
        valid_mask = data.notna()

        # ── Step 6: half-minimum imputation (per protein, Perseus-style) ──────
        arr = data.to_numpy(dtype=float, copy=True)
        nan_mask = np.isnan(arr)
        if nan_mask.any():
            with np.errstate(all="ignore"):
                row_half_min = np.nanmin(arr, axis=1, keepdims=True) / 2.0
            row_half_min = np.where(np.isnan(row_half_min), 0.0, row_half_min)
            arr[nan_mask] = np.broadcast_to(row_half_min, arr.shape)[nan_mask]
            data = pd.DataFrame(arr, index=data.index, columns=data.columns)

        qc_summary = {
            "proteins_input":        orig_proteins,
            "contaminants_removed":  n_contam,
            "proteins_after_qc":     len(data),
            "proteins_removed":      orig_proteins - len(data),
            "samples_input":         orig_samples,
            "samples_after_qc":      len(data.columns),
            "samples_removed":       orig_samples - len(data.columns),
            "log2_transformed":      log_done,
            "normalised":            norm_done,
            "normalisation_method":  "median_centering" if norm_done else "none",
            "irs_normalised":        irs_done,
            "missing_threshold":     missing_threshold,
            "imputation_method":     "half-minimum per protein",
        }
        return data, qc_summary, valid_mask

    # ── Supervised: two-group differential expression ─────────────────────────

    @staticmethod
    def _resolve_group_columns(
        requested: List[str],
        available: List[str],
    ) -> List[str]:
        """Resolve requested group columns against available columns robustly.

        Matching order:
        1) exact match
        2) case-insensitive trimmed match
        3) pandas-style duplicate suffix match (e.g. A -> A.1)
        """
        if not requested or not available:
            return []

        out: List[str] = []
        avail_map = {str(c).strip().lower(): c for c in available}
        avail_set = set(available)

        for raw in requested:
            key = str(raw).strip()
            if not key:
                continue

            if key in avail_set:
                out.append(key)
                continue

            low = key.lower()
            if low in avail_map:
                out.append(avail_map[low])
                continue

            # Common proteomics matrix headers: "A Intensity", "A_LFQ", etc.
            token_hits = [
                c for c in available
                if str(c).startswith(f"{key} ")
                or str(c).startswith(f"{key}_")
                or str(c).strip().lower().startswith(f"{low} ")
                or str(c).strip().lower().startswith(f"{low}_")
            ]
            if token_hits:
                out.extend(token_hits)
                continue

            # Handle duplicate-column suffixes produced by pandas (A, A.1, A.2).
            prefix = f"{key}."
            prefix_low = f"{low}."
            suffix_hits = [
                c for c in available
                if str(c).startswith(prefix) or str(c).strip().lower().startswith(prefix_low)
            ]
            if suffix_hits:
                out.extend(suffix_hits)

        # Preserve order and remove accidental duplicates.
        return [c for c in dict.fromkeys(out)]

    def _supervised(
        self,
        data: pd.DataFrame,
        valid_mask: pd.DataFrame,
        g1_cols: List[str],
        g2_cols: List[str],
        g1_label: str,
        g2_label: str,
        log2fc_cutoff: float = 1.0,
        adj_pval_cutoff: float = 0.05,
        missing_threshold: float = 0.5,
    ) -> pd.DataFrame:
        available_cols = list(data.columns)
        g1 = self._resolve_group_columns(g1_cols, available_cols)
        g2 = self._resolve_group_columns(g2_cols, available_cols)

        if not g1 or not g2:
            raise ValueError(
                f"Group columns not found in QC-filtered data.\n"
                f"Group1 requested: {g1_cols}\n"
                f"Group2 requested: {g2_cols}\n"
                f"Available columns: {list(data.columns)}"
            )

        if len(g1) < 2 or len(g2) < 2:
            raise ValueError(
                f"Differential analysis requires ≥2 samples per group. "
                f"Got: {g1_label}={len(g1)}, {g2_label}={len(g2)}. "
                f"Assign more samples to each group or use unsupervised mode."
            )

        # Group-aware missing value filter
        # Keep proteins detected in ≥threshold of samples in AT LEAST one group.
        # This is more sensitive than a global filter — a protein present in 100%
        # of disease samples but 0% of controls is biologically meaningful.
        vm = valid_mask.reindex(data.index, fill_value=True)
        valid_g1 = vm[g1].mean(axis=1) >= missing_threshold if g1 else pd.Series(True, index=data.index)
        valid_g2 = vm[g2].mean(axis=1) >= missing_threshold if g2 else pd.Series(True, index=data.index)
        data = data[valid_g1 | valid_g2]
        vm   = vm[valid_g1 | valid_g2]

        rows = []
        for protein in data.index:
            v1 = data.loc[protein, g1].values.astype(float)
            v2 = data.loc[protein, g2].values.astype(float)

            # Detection rates (from pre-imputation valid mask)
            if protein in vm.index:
                det_g1 = round(float(vm.loc[protein, g1].mean()), 3)
                det_g2 = round(float(vm.loc[protein, g2].mean()), 3)
            else:
                det_g1 = det_g2 = 1.0

            if len(v1) < 2 or len(v2) < 2:
                continue

            m1, m2 = v1.mean(), v2.mean()

            try:
                _, pval = stats.ttest_ind(v1, v2, equal_var=False)
            except Exception:
                continue

            if np.isnan(pval):
                continue

            # log2FC convention: log2(group2 / group1) — positive means up in
            # the second-named group (the "test" condition by convention).
            # Data is log2-space so we just subtract; cap at ±20.
            log2fc = max(-20.0, min(20.0, float(m2 - m1)))

            # Cohen's d (pooled-SD effect size)
            n1, n2 = len(v1), len(v2)
            s1, s2 = float(v1.std(ddof=1)), float(v2.std(ddof=1))
            sp = np.sqrt(((n1 - 1) * s1**2 + (n2 - 1) * s2**2) / max(n1 + n2 - 2, 1))
            cohens_d = float((m1 - m2) / sp) if sp > 0 else 0.0

            rows.append({
                "protein":                    protein,
                f"mean_{g1_label}":           round(float(m1), 4),
                f"mean_{g2_label}":           round(float(m2), 4),
                "log2_fold_change":           round(log2fc, 4),
                "cohens_d":                   round(cohens_d, 4),
                "p_value":                    float(pval),
                f"detection_{g1_label}":      det_g1,
                f"detection_{g2_label}":      det_g2,
                f"n_{g1_label}":              n1,
                f"n_{g2_label}":              n2,
            })

        if not rows:
            raise ValueError("No proteins had enough values in both groups for t-test.")

        df = pd.DataFrame(rows)

        # BH FDR correction
        _, adj_pvals, _, _ = multipletests(df["p_value"].values, method="fdr_bh")
        df["adj_p_value"] = adj_pvals
        df = self._add_significance(df, "log2_fold_change", adj_pval_cutoff, log2fc_cutoff)
        df = df.sort_values("adj_p_value").reset_index(drop=True)
        df.insert(0, "rank", range(1, len(df) + 1))
        return df

    # ── Pooled / no-replicates: log2 fold-change only ────────────────────────

    def _fold_change_only(
        self,
        data: pd.DataFrame,
        valid_mask: pd.DataFrame,
        g1_cols: List[str],
        g2_cols: List[str],
        g1_label: str,
        g2_label: str,
        log2fc_cutoff: float = 1.0,
        missing_threshold: float = 0.5,
    ) -> pd.DataFrame:
        """
        Pooled-design ranking — log2 fold-change only, no statistical test.

        Use when each "group" contains samples that aren't true biological
        replicates (e.g. one mouse per tissue, single MaxQuant pool per label).
        A Welch t-test on such data is statistically invalid; this branch
        reports log2(group2 / group1) and ranks proteins by |log2FC| only.
        """
        available_cols = list(data.columns)
        g1 = self._resolve_group_columns(g1_cols, available_cols)
        g2 = self._resolve_group_columns(g2_cols, available_cols)
        if not g1 or not g2:
            raise ValueError(
                "Group columns not found in QC-filtered data.\n"
                f"Group1 requested: {g1_cols}\n"
                f"Group1 matched: {g1}\n"
                f"Group2 requested: {g2_cols}\n"
                f"Group2 matched: {g2}\n"
                f"Available columns: {available_cols}"
            )

        vm = valid_mask.reindex(data.index, fill_value=True)
        valid_g1 = vm[g1].mean(axis=1) >= missing_threshold
        valid_g2 = vm[g2].mean(axis=1) >= missing_threshold
        data = data[valid_g1 | valid_g2]
        vm   = vm.loc[data.index]

        rows = []
        for protein in data.index:
            v1 = data.loc[protein, g1].values.astype(float)
            v2 = data.loc[protein, g2].values.astype(float)
            m1, m2 = float(np.nanmean(v1)), float(np.nanmean(v2))
            log2fc = max(-20.0, min(20.0, m2 - m1))    # log2(group2/group1)
            det1 = round(float(vm.loc[protein, g1].mean()), 3) if protein in vm.index else 1.0
            det2 = round(float(vm.loc[protein, g2].mean()), 3) if protein in vm.index else 1.0
            rows.append({
                "protein":               protein,
                f"mean_{g1_label}":      round(m1, 4),
                f"mean_{g2_label}":      round(m2, 4),
                "log2_fold_change":      round(log2fc, 4),
                "abs_log2_fold_change":  round(abs(log2fc), 4),
                f"detection_{g1_label}": det1,
                f"detection_{g2_label}": det2,
                f"n_{g1_label}":         len(g1),
                f"n_{g2_label}":         len(g2),
            })

        if not rows:
            raise ValueError("No proteins had enough values for fold-change ranking.")

        df = pd.DataFrame(rows)
        df["significance"] = np.where(
            df["log2_fold_change"].abs() >= log2fc_cutoff, "Notable", "NS",
        )
        df = df.sort_values("abs_log2_fold_change", ascending=False).reset_index(drop=True)
        df.insert(0, "rank", range(1, len(df) + 1))
        return df

    def _finalise_pooled(
        self,
        results_df, top_df, qc_summary, output_dir, file_name,
        g1_label, g2_label, n_sig,
        log2fc_cutoff, adj_pval_cutoff,
        data_path, sample_columns,
        group1_samples, group2_samples,
        missing_threshold, top_n,
    ) -> Dict[str, Any]:
        """Excel + code export for the fold-change-only branch."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = Path(file_name).stem.replace(" ", "_")
        excel_path = str(Path(output_dir) /
                          f"biomarkers_{safe_name}_pooled_{ts}.xlsx")
        # Re-use the regular Excel exporter — it handles missing adj_p_value gracefully
        results_df_export = results_df.copy()
        if "adj_p_value" not in results_df_export.columns:
            results_df_export["adj_p_value"] = pd.NA
            results_df_export["p_value"]     = pd.NA
        self._export_excel(
            results_df_export, top_df, qc_summary, excel_path,
            g1_label, g2_label, "supervised",
            adj_pval_cutoff, log2fc_cutoff, file_name,
            test_method="fold_change_only",
        )
        code = self._generate_code(
            data_path, sample_columns, group1_samples, group2_samples,
            g1_label, g2_label, "supervised", "generic",
            adj_pval_cutoff, log2fc_cutoff, missing_threshold, top_n,
            output_dir, file_name,
            pooled=True,
        )
        return {
            "omic_type":      self.omic_type,
            "top_biomarkers": top_df.to_dict("records"),
            "n_significant":  n_sig,
            "excel_path":     excel_path,
            "qc_summary":     qc_summary,
            "analysis_code":  code,
            "error":          None,
        }

    # ── Unsupervised: variability ranking ─────────────────────────────────────

    def _unsupervised(
        self,
        data: pd.DataFrame,
        valid_mask: pd.DataFrame,
    ) -> pd.DataFrame:
        rows = []
        for protein in data.index:
            vals = data.loc[protein].values.astype(float)
            if len(vals) < 3:
                continue
            m   = vals.mean()
            sd  = vals.std(ddof=1)
            cv  = (sd / abs(m) * 100) if m != 0 else 0.0
            mad = float(np.median(np.abs(vals - np.median(vals))))
            iqr = float(np.percentile(vals, 75) - np.percentile(vals, 25))

            # Detection rate (fraction of samples with genuine measurements)
            if protein in valid_mask.index:
                det = round(float(valid_mask.loc[protein].mean()), 3)
            else:
                det = 1.0

            rows.append({
                "protein":              protein,
                "mean_expression":      round(float(m), 4),
                "std_expression":       round(float(sd), 4),
                "cv_percent":           round(cv, 2),
                "median_abs_deviation": round(mad, 4),
                "iqr":                  round(iqr, 4),
                "detection_rate":       det,
                "n_samples":            len(vals),
                "significance":         "Top Variable",
            })

        df = pd.DataFrame(rows)
        df = df.sort_values("cv_percent", ascending=False).reset_index(drop=True)
        df.insert(0, "rank", range(1, len(df) + 1))
        return df

    # ── Test method selection ─────────────────────────────────────────────────

    @staticmethod
    def _resolve_test_method(
        test_method: str,
        is_paired: bool,
        g1_cols: List[str],
        g2_cols: List[str],
        data: pd.DataFrame,
    ) -> str:
        if is_paired:
            return "paired_t"
        if test_method and test_method != "auto":
            return test_method
        n1 = len([c for c in g1_cols if c in data.columns])
        n2 = len([c for c in g2_cols if c in data.columns])
        return "limma" if min(n1, n2) <= 4 else "welch"

    # ── eBayes prior estimation (Smyth 2004) ──────────────────────────────────

    @staticmethod
    def _fit_prior_variance(
        s2: np.ndarray, df: np.ndarray
    ) -> Tuple[float, float]:
        """
        Fit scaled inverse chi-squared prior on variances by moment matching.
        Returns (d0, s0_sq) — prior degrees of freedom and prior variance.
        """
        mask = (s2 > 0) & np.isfinite(s2) & np.isfinite(df) & (df > 0)
        s2c, dfc = s2[mask], df[mask]
        if len(s2c) < 3:
            return 4.0, float(np.nanmedian(s2))

        # Bias-corrected log-variance (removes chi-squared mean shift)
        z = np.log(s2c) - digamma(dfc / 2.0) + np.log(dfc / 2.0)
        s0_sq = float(np.exp(np.nanmean(z)))
        if not (np.isfinite(s0_sq) and s0_sq > 0):
            s0_sq = float(np.nanmedian(s2c))

        # Var(z) = ψ'(d0/2) + mean(ψ'(df/2)); solve for d0
        var_z    = float(np.nanvar(z, ddof=1))
        trig_df  = float(np.nanmean(polygamma(1, dfc / 2.0)))
        target   = var_z - trig_df
        if target <= 1e-8:
            return 20.0, s0_sq

        try:
            eq = lambda d0: float(polygamma(1, max(d0, 1e-6) / 2.0)) - target
            if eq(0.01) * eq(2000.0) >= 0:
                d0 = max(2.0 / target, 0.5)
            else:
                d0 = brentq(eq, 0.01, 2000.0, xtol=1e-4, maxiter=100)
        except Exception:
            d0 = max(2.0 / target, 0.5)

        return float(max(d0, 0.1)), float(s0_sq)

    # ── Limma moderated t-test ────────────────────────────────────────────────

    def _limma_ebayes(
        self,
        data: pd.DataFrame,
        valid_mask: pd.DataFrame,
        g1_cols: List[str],
        g2_cols: List[str],
        g1_label: str,
        g2_label: str,
        log2fc_cutoff: float = 1.0,
        adj_pval_cutoff: float = 0.05,
        missing_threshold: float = 0.5,
    ) -> pd.DataFrame:
        """
        Empirical Bayes moderated t-test (Smyth 2004).
        Shrinks per-protein variances towards a common prior — improves power
        when n is small (2–4 per group) by borrowing strength across proteins.
        """
        g1 = [c for c in g1_cols if c in data.columns]
        g2 = [c for c in g2_cols if c in data.columns]
        if not g1 or not g2:
            raise ValueError("Group columns not found after QC filtering.")
        if len(g1) < 2 or len(g2) < 2:
            raise ValueError(
                f"Limma eBayes requires ≥2 samples per group. Got {g1_label}={len(g1)}, {g2_label}={len(g2)}."
            )

        vm = valid_mask.reindex(data.index, fill_value=True)
        valid_g1 = vm[g1].mean(axis=1) >= missing_threshold
        valid_g2 = vm[g2].mean(axis=1) >= missing_threshold
        data = data[valid_g1 | valid_g2]
        vm   = vm.loc[data.index]

        per_prot = []
        for protein in data.index:
            v1 = data.loc[protein, g1].values.astype(float)
            v2 = data.loc[protein, g2].values.astype(float)
            if len(v1) < 2 or len(v2) < 2:
                continue
            n1, n2  = len(v1), len(v2)
            m1, m2  = v1.mean(), v2.mean()
            df_p    = n1 + n2 - 2
            s2_pool = ((n1 - 1) * v1.var(ddof=1) + (n2 - 1) * v2.var(ddof=1)) / max(df_p, 1)
            det1 = float(vm.loc[protein, g1].mean()) if protein in vm.index else 1.0
            det2 = float(vm.loc[protein, g2].mean()) if protein in vm.index else 1.0
            per_prot.append(dict(protein=protein, m1=m1, m2=m2, n1=n1, n2=n2,
                                 s2=s2_pool, df=df_p, det1=det1, det2=det2))

        if len(per_prot) < 3:
            raise ValueError("Fewer than 3 proteins had sufficient data for limma eBayes.")

        s2_arr = np.array([p["s2"] for p in per_prot])
        df_arr = np.array([p["df"] for p in per_prot], dtype=float)
        d0, s0_sq = self._fit_prior_variance(s2_arr, df_arr)
        logger.info("eBayes prior: d0=%.2f  s0²=%.4f  (shrinks towards common variance)", d0, s0_sq)

        rows = []
        for p in per_prot:
            df_p   = p["df"]
            s2_pos = (d0 * s0_sq + df_p * p["s2"]) / (d0 + df_p)
            df_pos = d0 + df_p
            se     = np.sqrt(s2_pos * (1.0 / p["n1"] + 1.0 / p["n2"]))
            if se < 1e-10:
                continue
            # Contrast direction: group2 − group1 (positive = up in group2)
            t_mod  = (p["m2"] - p["m1"]) / se
            pval   = 2.0 * float(stats.t.sf(abs(t_mod), df=df_pos))
            if np.isnan(pval):
                continue
            log2fc  = max(-20.0, min(20.0, float(p["m2"] - p["m1"])))
            sp      = np.sqrt(p["s2"]) if p["s2"] > 0 else 1e-10
            rows.append({
                "protein":                   p["protein"],
                f"mean_{g1_label}":          round(p["m1"], 4),
                f"mean_{g2_label}":          round(p["m2"], 4),
                "log2_fold_change":          round(log2fc, 4),
                "cohens_d":                  round(float((p["m2"] - p["m1"]) / sp), 4),
                "p_value":                   pval,
                "t_statistic_moderated":     round(t_mod, 4),
                "df_moderated":              round(df_pos, 1),
                "prior_df":                  round(d0, 1),
                f"detection_{g1_label}":     round(p["det1"], 3),
                f"detection_{g2_label}":     round(p["det2"], 3),
                f"n_{g1_label}":             p["n1"],
                f"n_{g2_label}":             p["n2"],
            })

        if not rows:
            raise ValueError("No proteins survived limma eBayes analysis.")
        df_res = pd.DataFrame(rows)
        _, adj_p, _, _ = multipletests(df_res["p_value"].values, method="fdr_bh")
        df_res["adj_p_value"] = adj_p
        df_res = self._add_significance(df_res, "log2_fold_change", adj_pval_cutoff, log2fc_cutoff)
        df_res = df_res.sort_values("adj_p_value").reset_index(drop=True)
        df_res.insert(0, "rank", range(1, len(df_res) + 1))
        return df_res

    # ── Paired t-test ─────────────────────────────────────────────────────────

    def _paired_ttest(
        self,
        data: pd.DataFrame,
        valid_mask: pd.DataFrame,
        g1_cols: List[str],
        g2_cols: List[str],
        g1_label: str,
        g2_label: str,
        log2fc_cutoff: float = 1.0,
        adj_pval_cutoff: float = 0.05,
        missing_threshold: float = 0.5,
    ) -> pd.DataFrame:
        """
        Paired t-test — g1_cols[i] must correspond to g2_cols[i] (same subject).
        Uses one-sample t-test on pairwise differences; handles NaN pairs by skipping them.
        Effect size: Cohen's d_z = mean(diff) / SD(diff).
        """
        g1 = [c for c in g1_cols if c in data.columns]
        g2 = [c for c in g2_cols if c in data.columns]
        n_pairs = min(len(g1), len(g2))
        if n_pairs < 2:
            raise ValueError(f"Paired t-test requires ≥2 pairs. Got {n_pairs}.")
        g1_p, g2_p = g1[:n_pairs], g2[:n_pairs]

        vm = valid_mask.reindex(data.index, fill_value=True)
        valid_g1 = vm[g1].mean(axis=1) >= missing_threshold
        valid_g2 = vm[g2].mean(axis=1) >= missing_threshold
        data = data[valid_g1 | valid_g2]
        vm   = vm.loc[data.index]

        rows = []
        for protein in data.index:
            v1   = data.loc[protein, g1_p].values.astype(float)
            v2   = data.loc[protein, g2_p].values.astype(float)
            diff = v2 - v1                       # group2 − group1 (positive = up in group2)
            ok   = ~np.isnan(diff)
            d    = diff[ok]
            if len(d) < 2:
                continue
            m1  = float(v1[~np.isnan(v1)].mean()) if not np.all(np.isnan(v1)) else 0.0
            m2  = float(v2[~np.isnan(v2)].mean()) if not np.all(np.isnan(v2)) else 0.0
            _, pval = stats.ttest_1samp(d, 0.0)
            if np.isnan(pval):
                continue
            log2fc   = max(-20.0, min(20.0, float(d.mean())))
            sd_diff  = float(d.std(ddof=1))
            cohens_dz = float(d.mean() / sd_diff) if sd_diff > 0 else 0.0
            det1 = round(float(vm.loc[protein, g1_p].mean()), 3) if protein in vm.index else 1.0
            det2 = round(float(vm.loc[protein, g2_p].mean()), 3) if protein in vm.index else 1.0
            rows.append({
                "protein":               protein,
                f"mean_{g1_label}":      round(m1, 4),
                f"mean_{g2_label}":      round(m2, 4),
                "log2_fold_change":      round(log2fc, 4),
                "cohens_dz":             round(cohens_dz, 4),
                "p_value":               float(pval),
                "n_pairs":               int(ok.sum()),
                f"detection_{g1_label}": det1,
                f"detection_{g2_label}": det2,
            })

        if not rows:
            raise ValueError("No proteins had sufficient paired observations.")
        df_res = pd.DataFrame(rows)
        _, adj_p, _, _ = multipletests(df_res["p_value"].values, method="fdr_bh")
        df_res["adj_p_value"] = adj_p
        df_res = self._add_significance(df_res, "log2_fold_change", adj_pval_cutoff, log2fc_cutoff)
        df_res = df_res.sort_values("adj_p_value").reset_index(drop=True)
        df_res.insert(0, "rank", range(1, len(df_res) + 1))
        return df_res

    # ── One-way ANOVA (>2 groups) ─────────────────────────────────────────────

    def _anova_multigroup(
        self,
        data: pd.DataFrame,
        valid_mask: pd.DataFrame,
        all_groups: Dict[str, List[str]],
        log2fc_cutoff: float = 1.0,
        adj_pval_cutoff: float = 0.05,
        missing_threshold: float = 0.5,
    ) -> pd.DataFrame:
        """
        One-way ANOVA across ≥2 groups with BH FDR correction.
        Effect size: η² (eta-squared).
        Post-hoc comparison: max pairwise log2FC between group means.
        """
        grp = {name: [c for c in cols if c in data.columns]
               for name, cols in all_groups.items()}
        grp = {k: v for k, v in grp.items() if len(v) >= 2}
        if len(grp) < 2:
            raise ValueError(
                f"ANOVA requires ≥2 groups with ≥2 samples each. Valid groups: {list(grp.keys())}"
            )

        vm = valid_mask.reindex(data.index, fill_value=True)
        keep = pd.Series(False, index=data.index)
        for cols in grp.values():
            keep |= (vm[cols].mean(axis=1) >= missing_threshold)
        data = data[keep]
        vm   = vm.loc[data.index]

        rows = []
        for protein in data.index:
            arrays, gnames, all_vals = [], [], []
            for gname, cols in grp.items():
                v = data.loc[protein, cols].values.astype(float)
                v = v[~np.isnan(v)]
                if len(v) >= 2:
                    arrays.append(v); gnames.append(gname); all_vals.extend(v.tolist())
            if len(arrays) < 2:
                continue
            try:
                f_stat, pval = stats.f_oneway(*arrays)
            except Exception:
                continue
            if np.isnan(pval):
                continue
            grand = np.mean(all_vals)
            ssb   = sum(len(v) * (v.mean() - grand) ** 2 for v in arrays)
            sst   = sum((x - grand) ** 2 for v in arrays for x in v)
            eta2  = ssb / sst if sst > 0 else 0.0
            means = [v.mean() for v in arrays]
            row = {
                "protein":     protein,
                "f_statistic": round(f_stat, 4),
                "p_value":     float(pval),
                "eta_squared": round(eta2, 4),
                "max_log2fc":  round(float(max(means) - min(means)), 4),
                "n_groups":    len(arrays),
            }
            for gn, v in zip(gnames, arrays):
                row[f"mean_{gn}"] = round(float(v.mean()), 4)
            rows.append(row)

        if not rows:
            raise ValueError("No proteins survived ANOVA analysis.")
        df_res = pd.DataFrame(rows)
        _, adj_p, _, _ = multipletests(df_res["p_value"].values, method="fdr_bh")
        df_res["adj_p_value"] = adj_p

        hi_pval    = min(0.01, adj_pval_cutoff / 5.0)
        trend_pval = adj_pval_cutoff * 2.0
        df_res["significance"] = "NS"
        hi  = (df_res["adj_p_value"] < hi_pval)        & (df_res["max_log2fc"].abs() >= log2fc_cutoff)
        sig = (df_res["adj_p_value"] < adj_pval_cutoff) & (df_res["max_log2fc"].abs() >= log2fc_cutoff)
        trn = (df_res["adj_p_value"] < trend_pval)      & ~sig
        df_res.loc[trn, "significance"] = "Trend"
        df_res.loc[sig, "significance"] = "Significant"
        df_res.loc[hi,  "significance"] = "Highly Significant"
        df_res = df_res.sort_values("adj_p_value").reset_index(drop=True)
        df_res.insert(0, "rank", range(1, len(df_res) + 1))
        return df_res

    # ── IRS normalisation (multi-batch TMT) ───────────────────────────────────

    @staticmethod
    def _apply_irs(
        data: pd.DataFrame,
        tmt_batches: Dict[str, Any],
    ) -> pd.DataFrame:
        """
        Internal Reference Scaling for multi-plex TMT (log2 data).
        For each plex: delta_protein = global_ref_protein − plex_ref_protein
        Add delta to all channels in that plex (log-space addition = ratio scaling).
        """
        batch_refs = {
            bname: binfo["reference"]
            for bname, binfo in tmt_batches.items()
            if binfo.get("reference") and binfo["reference"] in data.columns
        }
        if len(batch_refs) < 2:
            logger.warning("IRS: need ≥2 plexes with valid reference channels. Skipping.")
            return data

        ref_cols   = list(batch_refs.values())
        global_ref = data[ref_cols].mean(axis=1, skipna=True)   # per-protein mean across plexes

        data_irs = data.copy()
        for bname, ref_col in batch_refs.items():
            batch_samples = [
                c for c in tmt_batches[bname].get("samples", []) if c in data.columns
            ]
            if not batch_samples:
                continue
            delta = global_ref - data[ref_col]
            for col in batch_samples:
                data_irs[col] = data_irs[col].add(delta, axis=0)

        logger.info("IRS normalisation applied across %d plexes.", len(batch_refs))
        return data_irs

    # ── Shared significance labelling ─────────────────────────────────────────

    @staticmethod
    def _add_significance(
        df: pd.DataFrame,
        fc_col: str,
        adj_pval_cutoff: float,
        log2fc_cutoff: float,
    ) -> pd.DataFrame:
        hi_pval    = min(0.01, adj_pval_cutoff / 5.0)
        trend_pval = adj_pval_cutoff * 2.0
        df["significance"] = "NS"
        hi  = (df["adj_p_value"] < hi_pval)        & (df[fc_col].abs() >= log2fc_cutoff)
        sig = (df["adj_p_value"] < adj_pval_cutoff) & (df[fc_col].abs() >= log2fc_cutoff)
        trn = (df["adj_p_value"] < trend_pval)      & ~sig
        df.loc[trn, "significance"] = "Trend"
        df.loc[sig, "significance"] = "Significant"
        df.loc[hi,  "significance"] = "Highly Significant"
        return df

    # ── Code generation ───────────────────────────────────────────────────────

    @staticmethod
    def _generate_code(
        data_path, sample_columns, group1_samples, group2_samples,
        group1_label, group2_label, analysis_mode, data_type,
        adj_pval_cutoff, log2fc_cutoff, missing_threshold, top_n,
        output_dir, file_name,
        pooled: bool = False,
    ) -> str:
        """Return a self-contained, re-executable Python script for this analysis."""
        L: List[str] = []
        a = L.append

        a('#!/usr/bin/env python3')
        a('"""')
        a('Reproducible proteomics biomarker analysis')
        a('Auto-generated from the biomarker discovery platform.')
        a('Edit the Parameters section and re-run to reproduce or customise results.')
        a('"""')
        a('')
        a('import numpy as np')
        a('import pandas as pd')
        if analysis_mode == "supervised" and not pooled:
            a('from scipy import stats')
            a('from statsmodels.stats.multitest import multipletests')
        a('from pathlib import Path')
        a('')
        a('# ── Parameters ───────────────────────────────────────────────────────')
        a('DATA_PATH         = ' + repr(data_path))
        a('SAMPLE_COLUMNS    = ' + repr(sample_columns))
        if analysis_mode == "supervised":
            a('GROUP1_SAMPLES    = ' + repr(list(group1_samples or [])))
            a('GROUP2_SAMPLES    = ' + repr(list(group2_samples or [])))
            a('GROUP1_LABEL      = ' + repr(group1_label))
            a('GROUP2_LABEL      = ' + repr(group2_label))
        a('ANALYSIS_MODE     = ' + repr(analysis_mode))
        a('ADJ_PVAL_CUTOFF   = ' + str(adj_pval_cutoff))
        a('LOG2FC_CUTOFF     = ' + str(log2fc_cutoff))
        a('MISSING_THRESHOLD = ' + str(missing_threshold))
        a('TOP_N             = ' + str(max(int(top_n), 50)))
        a('OUTPUT_DIR        = ' + repr(output_dir))
        a('FILE_NAME         = ' + repr(file_name))
        a('')
        a('# ── 1. Load ──────────────────────────────────────────────────────────')
        a('df_raw = pd.read_csv(DATA_PATH, index_col=0)')
        a('avail  = [c for c in SAMPLE_COLUMNS if c in df_raw.columns]')
        a('data   = df_raw[avail].apply(pd.to_numeric, errors="coerce")')
        a('print(f"Loaded: {len(data)} proteins x {len(data.columns)} samples")')
        a('')
        a('# ── 2. QC: log2 transform, protein/sample filter ─────────────────────')
        a('if data.max().max() > 100:')
        a('    data = np.log2(data.replace(0, np.nan) + 1)')
        a('    print("Log2 transform applied")')
        a('miss_prot = data.isna().mean(axis=1)')
        a('data = data.loc[miss_prot <= MISSING_THRESHOLD]')
        a('miss_samp = data.isna().mean(axis=0)')
        a('data = data.loc[:, miss_samp <= 0.80]')
        a('')
        a('# ── 3. Normalisation: median-centering per sample ───────────────────')
        a('sample_medians = data.median(axis=0, skipna=True)')
        a('global_median  = float(sample_medians.median())')
        a('data = data.subtract(sample_medians, axis=1).add(global_median)')
        a('print("Median-centering normalisation applied")')
        a('')
        a('# ── 4. Save valid mask, then impute missing values ───────────────────')
        a('valid_mask = data.notna()  # True = genuinely measured')
        a('arr = data.to_numpy(dtype=float, copy=True)')
        a('nan_mask = np.isnan(arr)')
        a('if nan_mask.any():')
        a('    row_min = np.nanmin(arr, axis=1, keepdims=True) / 2.0')
        a('    row_min = np.where(np.isnan(row_min), 0.0, row_min)')
        a('    arr[nan_mask] = np.broadcast_to(row_min, arr.shape)[nan_mask]')
        a('    data = pd.DataFrame(arr, index=data.index, columns=data.columns)')
        a('print(f"After QC: {len(data)} proteins, {len(data.columns)} samples")')
        a('')

        hi_pval    = min(0.01, adj_pval_cutoff / 5.0)
        trend_pval = adj_pval_cutoff * 2.0

        if pooled:
            a('# ── 5. POOLED DESIGN — log2 fold-change ranking, NO t-test ─────────')
            a('# Each group contains non-replicate samples (different tissues or')
            a('# single MaxQuant pools per condition). A Welch t-test would treat')
            a('# between-sample variation as biological noise, producing invalid')
            a('# p-values. We rank by |log2(group2/group1)| instead.')
            a('g1 = [c for c in GROUP1_SAMPLES if c in data.columns]')
            a('g2 = [c for c in GROUP2_SAMPLES if c in data.columns]')
            a('valid_g1 = valid_mask[g1].mean(axis=1) >= MISSING_THRESHOLD')
            a('valid_g2 = valid_mask[g2].mean(axis=1) >= MISSING_THRESHOLD')
            a('data = data[valid_g1 | valid_g2]')
            a('')
            a('rows = []')
            a('for protein in data.index:')
            a('    v1 = data.loc[protein, g1].values.astype(float)')
            a('    v2 = data.loc[protein, g2].values.astype(float)')
            a('    m1 = float(np.nanmean(v1))')
            a('    m2 = float(np.nanmean(v2))')
            a('    lfc = max(-20.0, min(20.0, m2 - m1))   # log2(group2/group1)')
            a('    rows.append({')
            a('        "protein":              protein,')
            a('        f"mean_{GROUP1_LABEL}": round(m1, 4),')
            a('        f"mean_{GROUP2_LABEL}": round(m2, 4),')
            a('        "log2_fold_change":     round(lfc, 4),')
            a('        "abs_log2_fold_change": round(abs(lfc), 4),')
            a('    })')
            a('df = pd.DataFrame(rows)')
            a('df["significance"] = np.where(df["log2_fold_change"].abs() >= LOG2FC_CUTOFF, "Notable", "NS")')
            a('df = df.sort_values("abs_log2_fold_change", ascending=False).reset_index(drop=True)')
            a('df.insert(0, "rank", range(1, len(df) + 1))')
            a('n_sig = (df["log2_fold_change"].abs() >= LOG2FC_CUTOFF).sum()')
            a(f'print(f"Proteins with |log2FC|>={log2fc_cutoff}: {{n_sig}}")')
        elif analysis_mode == "supervised":
            a('# ── 5. Group-aware filter + Welch t-test + BH FDR ──────────────────')
            a('g1 = [c for c in GROUP1_SAMPLES if c in data.columns]')
            a('g2 = [c for c in GROUP2_SAMPLES if c in data.columns]')
            a('# Keep proteins detected in >=MISSING_THRESHOLD of samples in at least one group')
            a('valid_g1 = valid_mask[g1].mean(axis=1) >= MISSING_THRESHOLD')
            a('valid_g2 = valid_mask[g2].mean(axis=1) >= MISSING_THRESHOLD')
            a('data = data[valid_g1 | valid_g2]')
            a('valid_mask = valid_mask.loc[data.index]')
            a('')
            a('rows = []')
            a('for protein in data.index:')
            a('    v1 = data.loc[protein, g1].values.astype(float)')
            a('    v2 = data.loc[protein, g2].values.astype(float)')
            a('    if len(v1) < 2 or len(v2) < 2:')
            a('        continue')
            a('    m1, m2 = v1.mean(), v2.mean()')
            a('    _, pval = stats.ttest_ind(v1, v2, equal_var=False)')
            a('    if np.isnan(pval):')
            a('        continue')
            a('    # log2FC convention: log2(group2/group1) — positive means up in group2.')
            a('    # Data is log2-space so we just subtract; cap at ±20.')
            a('    lfc = max(-20.0, min(20.0, float(m2 - m1)))')
            a('    # Cohen\'s d (pooled-SD effect size) — same convention')
            a('    n1, n2 = len(v1), len(v2)')
            a('    sp = np.sqrt(((n1-1)*v1.std(ddof=1)**2 + (n2-1)*v2.std(ddof=1)**2) / max(n1+n2-2, 1))')
            a('    cohens_d = float((m2 - m1) / sp) if sp > 0 else 0.0')
            a('    det_g1 = round(float(valid_mask.loc[protein, g1].mean()), 3)')
            a('    det_g2 = round(float(valid_mask.loc[protein, g2].mean()), 3)')
            a('    rows.append({')
            a('        "protein":                        protein,')
            a('        f"mean_{GROUP1_LABEL}":           round(float(m1), 4),')
            a('        f"mean_{GROUP2_LABEL}":           round(float(m2), 4),')
            a('        "log2_fold_change":               round(lfc, 4),')
            a('        "cohens_d":                       round(cohens_d, 4),')
            a('        "p_value":                        float(pval),')
            a('        f"detection_{GROUP1_LABEL}":      det_g1,')
            a('        f"detection_{GROUP2_LABEL}":      det_g2,')
            a('    })')
            a('df = pd.DataFrame(rows)')
            a('_, adj_p, _, _ = multipletests(df["p_value"].values, method="fdr_bh")')
            a('df["adj_p_value"] = adj_p')
            a('df["significance"] = "NS"')
            a(f'hi  = (df["adj_p_value"] < {hi_pval})      & (df["log2_fold_change"].abs() >= LOG2FC_CUTOFF)')
            a(f'sig = (df["adj_p_value"] < ADJ_PVAL_CUTOFF) & (df["log2_fold_change"].abs() >= LOG2FC_CUTOFF)')
            a(f'trn = (df["adj_p_value"] < {trend_pval})   & ~sig')
            a('df.loc[trn, "significance"]  = "Trend"')
            a('df.loc[sig, "significance"]  = "Significant"')
            a('df.loc[hi,  "significance"]  = "Highly Significant"')
            a('df = df.sort_values("adj_p_value").reset_index(drop=True)')
            a('df.insert(0, "rank", range(1, len(df) + 1))')
            a('n_sig = ((df["adj_p_value"] < ADJ_PVAL_CUTOFF) &')
            a('         (df["log2_fold_change"].abs() >= LOG2FC_CUTOFF)).sum()')
            a(f'print(f"Significant (adj.p<{adj_pval_cutoff}, |log2FC|>={log2fc_cutoff}): {{n_sig}}")')
        else:
            a('# ── 5. Unsupervised: CV / MAD / IQR ranking ─────────────────────────')
            a('rows = []')
            a('for protein in data.index:')
            a('    vals = data.loc[protein].values.astype(float)')
            a('    if len(vals) < 3:')
            a('        continue')
            a('    m   = vals.mean()')
            a('    sd  = vals.std(ddof=1)')
            a('    cv  = (sd / abs(m) * 100) if m != 0 else 0.0')
            a('    mad = float(np.median(np.abs(vals - np.median(vals))))')
            a('    iqr = float(np.percentile(vals, 75) - np.percentile(vals, 25))')
            a('    det = round(float(valid_mask.loc[protein].mean()), 3) if protein in valid_mask.index else 1.0')
            a('    rows.append({"protein": protein, "mean_expression": round(float(m), 4),')
            a('                 "std_expression": round(float(sd), 4), "cv_percent": round(cv, 2),')
            a('                 "median_abs_deviation": round(mad, 4), "iqr": round(iqr, 4),')
            a('                 "detection_rate": det})')
            a('df = pd.DataFrame(rows).sort_values("cv_percent", ascending=False).reset_index(drop=True)')
            a('df.insert(0, "rank", range(1, len(df) + 1))')

        a('')
        a('# ── 6. Save results ──────────────────────────────────────────────────')
        a('Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)')
        a('out = str(Path(OUTPUT_DIR) / f"{FILE_NAME}_reproduced_results.csv")')
        a('df.head(TOP_N).to_csv(out, index=False)')
        a('print(f"Saved -> {out}")')
        a('print(df.head(10).to_string(index=False))')

        return '\n'.join(L)

    # ── Protein-metadata enrichment ───────────────────────────────────────────

    @staticmethod
    def _build_protein_metadata(
        df_raw: pd.DataFrame, sample_cols: List[str]
    ) -> pd.DataFrame:
        """
        Build a DataFrame indexed by accession with `protein_name` and
        `gene_name` columns, sourced ONLY from the user's uploaded file.

        We never want the LLM to guess identities from accession numbers —
        every result row is enriched with the actual values that the user
        supplied, so the chatbot's biological interpretation is grounded.
        """
        # All non-sample columns are candidates for identifier metadata.
        meta_cols = [c for c in df_raw.columns if c not in set(sample_cols)]
        meta = df_raw[meta_cols].copy() if meta_cols else pd.DataFrame(index=df_raw.index)

        def _pick(hints: Tuple[str, ...]) -> Optional[str]:
            for c in meta.columns:
                lc = str(c).strip().lower().replace(" ", "").replace("_", "").replace("-", "")
                for h in hints:
                    if h.replace(" ", "").replace("_", "").replace("-", "") in lc:
                        return c
            return None

        name_col = _pick(("proteinname", "proteinnames", "description", "protein"))
        gene_col = _pick(("genename", "genenames", "genesymbol", "gene"))

        out = pd.DataFrame(index=df_raw.index)
        out["protein_name"] = meta[name_col].astype(str) if name_col else ""
        out["gene_name"]    = meta[gene_col].astype(str) if gene_col else ""
        return out

    @staticmethod
    def _enrich_with_metadata(
        results_df: pd.DataFrame, meta_lookup: pd.DataFrame,
    ) -> pd.DataFrame:
        """Attach protein_name + gene_name columns to a results DataFrame
        whose `protein` column holds the accession (the proteins_df row index)."""
        if results_df.empty or "protein" not in results_df.columns:
            return results_df
        # Map accession → metadata via the index of meta_lookup
        names = meta_lookup["protein_name"].reindex(results_df["protein"]).values
        genes = meta_lookup["gene_name"].reindex(results_df["protein"]).values
        # Insert near the front so they show up first in CSV / Excel previews
        results_df = results_df.copy()
        if "protein_name" not in results_df.columns:
            results_df.insert(
                results_df.columns.get_loc("protein") + 1,
                "protein_name", names,
            )
        if "gene_name" not in results_df.columns:
            results_df.insert(
                results_df.columns.get_loc("protein_name") + 1,
                "gene_name", genes,
            )
        return results_df

    # ── Excel export ──────────────────────────────────────────────────────────

    def _export_excel(
        self,
        all_results: pd.DataFrame,
        top_results: pd.DataFrame,
        qc_summary: Dict,
        output_path: str,
        g1_label: str,
        g2_label: str,
        analysis_mode: str,
        adj_pval_cutoff: float,
        log2fc_cutoff: float,
        file_name: str,
        test_method: str = "welch",
    ) -> None:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            top_results.to_excel(writer, sheet_name="Top Biomarkers", index=False)
            all_results.to_excel(writer, sheet_name="All Results",    index=False)

            qc_df = pd.DataFrame(
                [{"Parameter": k.replace("_", " ").title(), "Value": v}
                 for k, v in qc_summary.items()]
            )
            qc_df.to_excel(writer, sheet_name="QC Summary", index=False)

            params = {
                "Analysis Date":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "File":               file_name,
                "Mode":               analysis_mode.capitalize(),
                "Group 1":            g1_label if analysis_mode == "supervised" else "N/A",
                "Group 2":            g2_label if analysis_mode == "supervised" else "N/A",
                "Adj P-value Cut":    adj_pval_cutoff,
                "Log2 FC Cut":        log2fc_cutoff,
                "Statistical Test":   {
                    "welch":    "Welch two-sample t-test",
                    "limma":    "Limma moderated t-test (eBayes)",
                    "paired_t": "Paired t-test",
                    "anova":    "One-way ANOVA",
                }.get(test_method, "CV ranking") if analysis_mode == "supervised" else "CV ranking",
                "FDR Method":         ("Benjamini-Hochberg" if analysis_mode == "supervised"
                                       else "N/A"),
                "Normalisation":      "Median centering per sample",
                "Imputation":         "Half-minimum per protein",
                "Effect Size":        "Cohen's d (pooled SD)" if analysis_mode == "supervised" else "N/A",
            }
            params_df = pd.DataFrame(
                [{"Parameter": k, "Value": v} for k, v in params.items()]
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

        self._format_workbook(output_path, analysis_mode)

    def _format_workbook(self, path: str, analysis_mode: str) -> None:
        wb = openpyxl.load_workbook(path)
        hdr_fill, hdr_font = _header_style()
        border = _thin_border()

        sig_fills = {
            "Highly Significant": PatternFill(start_color=_HIGHLY_SIG,  end_color=_HIGHLY_SIG,  fill_type="solid"),
            "Significant":        PatternFill(start_color=_SIGNIFICANT, end_color=_SIGNIFICANT, fill_type="solid"),
            "Trend":              PatternFill(start_color=_TREND,       end_color=_TREND,       fill_type="solid"),
            "Top Variable":       PatternFill(start_color=_VARIABLE,    end_color=_VARIABLE,    fill_type="solid"),
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for cell in ws[1]:
                cell.fill      = hdr_fill
                cell.font      = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = border
            ws.row_dimensions[1].height = 28

            for col in ws.columns:
                letter  = get_column_letter(col[0].column)
                max_len = max(
                    (len(str(c.value)) for c in col if c.value is not None),
                    default=8,
                )
                ws.column_dimensions[letter].width = min(max_len + 4, 40)

            if sheet_name == "Top Biomarkers":
                sig_col_idx = next(
                    (i for i, c in enumerate(ws[1], 1)
                     if c.value and "significance" in str(c.value).lower()),
                    None,
                )
                for row in ws.iter_rows(min_row=2):
                    sig_val = (str(row[sig_col_idx - 1].value)
                               if sig_col_idx and row[sig_col_idx - 1].value
                               else "NS")
                    fill = sig_fills.get(sig_val)
                    for cell in row:
                        if fill:
                            cell.fill = fill
                        cell.border    = border
                        cell.alignment = Alignment(horizontal="center")

            ws.freeze_panes = "A2"

        # Legend sheet
        legend_ws = wb.create_sheet("Legend")
        legend_data = [
            ("Colour",      "Significance",         "Criteria"),
            ("Dark Green",  "Highly Significant",   "Adj p < 0.01 AND |log₂FC| ≥ threshold"),
            ("Light Green", "Significant",          "Adj p < threshold AND |log₂FC| ≥ threshold"),
            ("Yellow",      "Trend",                "Adj p < 2×threshold (FC not required)"),
            ("Pale Blue",   "Top Variable",         "Top CV% proteins (unsupervised mode)"),
            ("White",       "Not Significant (NS)", "Does not meet above thresholds"),
        ]
        colour_fills = [
            None,
            sig_fills["Highly Significant"],
            sig_fills["Significant"],
            sig_fills["Trend"],
            sig_fills["Top Variable"],
            None,
        ]

        for r_idx, (row_data, row_fill) in enumerate(zip(legend_data, colour_fills), 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = legend_ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                if r_idx == 1:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                elif row_fill:
                    cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left")

        for col in legend_ws.columns:
            letter = get_column_letter(col[0].column)
            legend_ws.column_dimensions[letter].width = 45

        wb.save(path)
