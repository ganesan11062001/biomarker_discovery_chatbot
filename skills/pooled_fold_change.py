"""
skills/pooled_fold_change.py
Analysis Layer — PooledFoldChangeSkill

Handles pooled proteomics designs where each group is a single pooled sample
(n=1).  Because t-tests require replicates, this skill computes log2 fold
changes across all pairwise contrasts derived from whatever groups are present
in the uploaded data.

Expected input
--------------
A two-sheet Excel file:
  • "Identifier Info"  — columns include "Majority protein IDs", label codes,
                         and a metadata row for group names (parsed by the
                         ingestion agent into a label_map dict).
  • "Proteins"         — rows = proteins, columns include:
                           "Majority protein IDs"
                           "<label> Spectral count <file>"   (SpC columns)
                           "<label> Intensity <file>"        (Intensity columns)

Contrasts analysed
------------------
  All pairwise comparisons built dynamically from whatever group labels are
  found in the data (e.g. ControlA_vs_ControlB, Disease_vs_Control, etc.).
  No hardcoded group names or contrast lists.

Outputs
-------
  • outputs/<stem>_cleaned_expression.csv
  • outputs/<stem>_fold_changes.csv
  • outputs/<stem>_top50_rescued.csv
  • outputs/<stem>_heatmap.png
  • outputs/<stem>_barchart_<contrast>.png   (one per contrast)
  • outputs/<stem>_scatter_<g1>_vs_<g2>.png
  • outputs/<stem>_venn_rescued.png
  • outputs/<stem>_biomarker_report.xlsx
"""
from __future__ import annotations

import logging
import re
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

from skills.base_skill import BaseOmicsSkill, OmicsAnalysisResult

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────

# Patterns that mark contaminant / reverse entries
_CONTAMINANT_PREFIXES = ("REV__", "CON__")
_CONTAMINANT_KEYWORDS = ("keratin", "krt", "bovine")

_LOG2FC_THRESHOLD = 1.0    # |log2FC| ≥ 1 → "changed"
_PSEUDOCOUNT      = 1.0    # added before log2 to avoid log(0)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _remove_contaminants(df: pd.DataFrame) -> pd.DataFrame:
    """Drop rows whose index matches contaminant / reverse-hit patterns."""
    mask = pd.Series(True, index=df.index)
    idx_lower = df.index.astype(str).str.lower()
    for prefix in _CONTAMINANT_PREFIXES:
        mask &= ~df.index.astype(str).str.startswith(prefix, na=False)
    for kw in _CONTAMINANT_KEYWORDS:
        mask &= ~idx_lower.str.contains(kw, na=False)
    removed = (~mask).sum()
    if removed:
        logger.info("Removed %d contaminant rows.", removed)
    return df[mask]


def _parse_proteins_sheet(
    raw_path: str,
    label_map: Dict[str, str],
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, str]]:
    """
    Parse the 'Proteins' sheet and return:
      spc_df   — SpC matrix  (proteins × group-labels)
      int_df   — Intensity matrix (proteins × group-labels)
      col_map  — { original_col_name: group_label }
    """
    xl = pd.ExcelFile(raw_path, engine="openpyxl")
    sheet = "Proteins" if "Proteins" in xl.sheet_names else xl.sheet_names[0]
    df = xl.parse(sheet)

    # Set protein-ID index — try multiple column-name patterns before fallback
    _ID_KEYWORDS = (
        "majority protein", "identified protein", "accession",
        "uniprot", "gene name", "protein id", "protein name",
    )
    id_col = next(
        (c for c in df.columns
         if any(kw in str(c).lower() for kw in _ID_KEYWORDS)),
        df.columns[0],
    )
    df = df.set_index(id_col)
    df.index = df.index.astype(str).str.strip()

    # Remove contaminants
    df = _remove_contaminants(df)

    # Identify SpC and Intensity columns per label
    spc_cols:  Dict[str, str] = {}
    int_cols:  Dict[str, str] = {}

    for col in df.columns:
        col_str = str(col).strip()
        col_lower = col_str.lower()
        for code, group in label_map.items():
            prefix = code.lower()
            is_match = (
                col_lower.startswith(f"{prefix} ") or
                col_lower.startswith(f"{prefix}_") or
                f" {prefix} " in col_lower
            )
            if not is_match:
                continue
            if "spectral count" in col_lower or "spc" in col_lower or "ms/ms count" in col_lower:
                if group not in spc_cols:
                    spc_cols[group] = col_str
            elif "intensity" in col_lower and "lfq" not in col_lower:
                if group not in int_cols:
                    int_cols[group] = col_str

    # Fall back: grab columns by position if pattern matching missed everything
    if not spc_cols and not int_cols:
        numeric_cols = [c for c in df.columns if pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.5]
        half = len(numeric_cols) // 2
        groups = list(label_map.values())
        for i, col in enumerate(numeric_cols):
            g = groups[i % len(groups)] if groups else str(i)
            if i < half:
                spc_cols[g] = col
            else:
                int_cols[g] = col

    def _build_matrix(col_dict: Dict[str, str]) -> pd.DataFrame:
        if not col_dict:
            return pd.DataFrame(index=df.index)
        mat = pd.DataFrame(
            {grp: pd.to_numeric(df[orig], errors="coerce").fillna(0)
             for grp, orig in col_dict.items()},
            index=df.index,
        )
        return mat

    spc_df = _build_matrix(spc_cols)
    int_df = _build_matrix(int_cols)

    col_map = {**{v: k for k, v in spc_cols.items()}, **{v: k for k, v in int_cols.items()}}
    return spc_df, int_df, col_map


def _log2_transform(mat: pd.DataFrame, pseudocount: float = _PSEUDOCOUNT) -> pd.DataFrame:
    return np.log2(mat + pseudocount)


def _compute_fold_changes(
    log2_mat: pd.DataFrame,
    contrasts: List[Tuple[str, str, str]],
) -> pd.DataFrame:
    """Return a DataFrame with one column per contrast (log2FC values)."""
    records: Dict[str, pd.Series] = {}
    for num, den, name in contrasts:
        if num in log2_mat.columns and den in log2_mat.columns:
            records[name] = log2_mat[num] - log2_mat[den]
        else:
            logger.warning("Contrast %s: missing column '%s' or '%s'.", name, num, den)
    return pd.DataFrame(records, index=log2_mat.index) if records else pd.DataFrame(index=log2_mat.index)


def _build_contrasts(
    label_map: Dict[str, str],
    available_groups: set,
) -> List[Tuple[str, str, str]]:
    """
    Build all pairwise contrasts from whatever groups are present in the data.
    Groups are sorted alphabetically so contrast order is deterministic.
    """
    groups = sorted(available_groups)
    pairwise: List[Tuple[str, str, str]] = []
    for i, g1 in enumerate(groups):
        for g2 in groups[i + 1:]:
            pairwise.append((g1, g2, f"{g1}_vs_{g2}"))
    return pairwise


def _top_n_by_variance(log2_mat: pd.DataFrame, n: int = 50) -> pd.DataFrame:
    var = log2_mat.var(axis=1)
    top_idx = var.nlargest(n).index
    return log2_mat.loc[top_idx]


def _rescue_score(fc_df: pd.DataFrame) -> pd.Series:
    """
    Activity score: sum of positive fold-changes across ALL contrasts.
    Ranks proteins that are consistently upregulated across the most comparisons.
    Fully generic — works for any label_map and any set of contrasts.
    """
    score = pd.Series(0.0, index=fc_df.index)
    for col in fc_df.columns:
        score += fc_df[col].clip(lower=0)
    return score


# ── Plotting ──────────────────────────────────────────────────────────────────

def _plot_heatmap(
    top_mat: pd.DataFrame, output_path: str, title: str = "Top 50 proteins by variance"
) -> str:
    """Seaborn clustermap; returns saved path."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import seaborn as sns

        g = sns.clustermap(
            top_mat,
            cmap="RdBu_r",
            center=0,
            figsize=(10, 14),
            yticklabels=True,
            xticklabels=True,
            linewidths=0.3,
        )
        g.fig.suptitle(title, y=1.01, fontsize=13)
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        plt.close("all")
        return output_path
    except Exception as exc:
        logger.warning("Heatmap generation failed: %s", exc)
        return ""


def _plot_barchart(
    fc_series: pd.Series,
    contrast_name: str,
    output_path: str,
    top_n: int = 20,
) -> str:
    """Bar chart of top N up- and down-regulated proteins."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        up   = fc_series.nlargest(top_n)
        down = fc_series.nsmallest(top_n)
        combined = pd.concat([up, down]).drop_duplicates().sort_values()

        colors = ["#d73027" if v > 0 else "#4575b4" for v in combined.values]
        fig, ax = plt.subplots(figsize=(10, max(6, len(combined) * 0.28)))
        combined.plot(kind="barh", ax=ax, color=colors, edgecolor="white", linewidth=0.4)
        ax.axvline(0, color="black", linewidth=0.8)
        ax.set_xlabel("log₂ Fold Change")
        ax.set_title(f"Top proteins — {contrast_name}")
        ax.tick_params(axis="y", labelsize=7)
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        plt.close("all")
        return output_path
    except Exception as exc:
        logger.warning("Bar chart failed (%s): %s", contrast_name, exc)
        return ""


def _plot_scatter(
    spc_df: pd.DataFrame, g1_col: str, g2_col: str, output_path: str
) -> str:
    """Scatter plot of one group vs another in log2 space.

    Parameters
    ----------
    g1_col, g2_col  Column names for the two groups being compared. The plot
                    is fully agnostic to what biological condition each holds.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        x = np.log2(spc_df[g1_col].clip(lower=0) + _PSEUDOCOUNT)
        y = np.log2(spc_df[g2_col].clip(lower=0) + _PSEUDOCOUNT)

        fig, ax = plt.subplots(figsize=(7, 7))
        ax.scatter(x, y, alpha=0.5, s=20, color="#555555")
        lims = [min(x.min(), y.min()) - 0.5, max(x.max(), y.max()) + 0.5]
        ax.plot(lims, lims, "r--", linewidth=0.8, label="y = x")
        ax.set_xlabel(f"log₂(SpC + 1)  {g1_col}")
        ax.set_ylabel(f"log₂(SpC + 1)  {g2_col}")
        ax.set_title(f"{g1_col} vs {g2_col}  SpC (log₂)")
        ax.legend(fontsize=9)
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        plt.close("all")
        return output_path
    except Exception as exc:
        logger.warning("Scatter plot failed: %s", exc)
        return ""


def _plot_venn(
    fc_df: pd.DataFrame, output_path: str, threshold: float = _LOG2FC_THRESHOLD
) -> str:
    """Venn diagram of proteins UP in the first two actual contrast columns."""
    if len(fc_df.columns) < 2:
        return ""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib_venn import venn2

        col_a, col_b = fc_df.columns[0], fc_df.columns[1]
        set_a = set(fc_df.index[fc_df[col_a] >= threshold])
        set_b = set(fc_df.index[fc_df[col_b] >= threshold])

        fig, ax = plt.subplots(figsize=(7, 5))
        venn2([set_a, set_b],
              set_labels=(f"UP: {col_a}", f"UP: {col_b}"), ax=ax)
        ax.set_title(f"Overlap of upregulated proteins\n{col_a}  |  {col_b}")
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        plt.close("all")
        return output_path
    except ImportError:
        logger.warning("matplotlib-venn not installed — Venn diagram skipped.")
        return ""
    except Exception as exc:
        logger.warning("Venn diagram failed: %s", exc)
        return ""


# ── Excel report ──────────────────────────────────────────────────────────────

def _write_excel_report(
    cleaned: pd.DataFrame,
    fc_df: pd.DataFrame,
    top50: pd.DataFrame,
    top_biomarkers: List[Dict[str, Any]],
    output_path: str,
) -> None:
    """Write a 4-sheet formatted Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF")
    up_fill      = PatternFill("solid", fgColor="FCE4D6")
    down_fill    = PatternFill("solid", fgColor="DDEEFF")
    center_align = Alignment(horizontal="center")

    def _write_df(ws, df: pd.DataFrame, *, include_index: bool = True) -> None:
        rows = list(dataframe_to_rows(df, index=include_index, header=True))
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                if r_idx == 1:
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = center_align

    # Sheet 1 — cleaned log2 expression
    ws1 = wb.active
    ws1.title = "Log2 Expression"
    _write_df(ws1, cleaned)

    # Sheet 2 — fold changes
    ws2 = wb.create_sheet("Fold Changes")
    _write_df(ws2, fc_df)
    # Colour cells
    for row in ws2.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= _LOG2FC_THRESHOLD:
                    cell.fill = up_fill
                elif cell.value <= -_LOG2FC_THRESHOLD:
                    cell.fill = down_fill

    # Sheet 3 — top 50 by variance
    ws3 = wb.create_sheet("Top50 by Variance")
    _write_df(ws3, top50)

    # Sheet 4 — top biomarkers ranked by rescue score
    ws4 = wb.create_sheet("Top Biomarkers")
    tb_df = pd.DataFrame(top_biomarkers)
    _write_df(ws4, tb_df, include_index=False)

    wb.save(output_path)


# ── Main skill ────────────────────────────────────────────────────────────────

class PooledFoldChangeSkill(BaseOmicsSkill):
    """
    Analysis skill for pooled proteomics designs (n=1 per group).

    Computes log2 fold changes across configurable contrasts without
    relying on replicates or statistical tests.
    """

    @property
    def omic_type(self) -> str:
        return "proteomics_pooled"

    def execute(self, **kwargs: Any) -> OmicsAnalysisResult:  # type: ignore[override]
        raw_path  = kwargs.get("raw_data_path") or kwargs.get("data_path", "")
        label_map = kwargs.get("label_map")
        output_dir = str(kwargs.get("output_dir", "outputs"))
        top_n     = int(kwargs.get("top_n", 50))
        file_name = kwargs.get("file_name") or Path(raw_path).stem

        if not label_map:
            raise ValueError(
                "label_map is required for pooled fold-change analysis but was not provided. "
                "The uploaded file must contain a metadata sheet with group code → group name "
                "mappings (e.g. A → Control, B → Disease)."
            )

        Path(output_dir).mkdir(parents=True, exist_ok=True)

        try:
            return self._run(raw_path, label_map, output_dir, top_n, file_name)
        except Exception as exc:
            logger.error("PooledFoldChangeSkill failed:\n%s", traceback.format_exc())
            return OmicsAnalysisResult(
                omic_type=self.omic_type,
                top_biomarkers=[],
                n_significant=0,
                excel_path=None,
                qc_summary={},
                error=str(exc),
            )

    # ── Core pipeline ─────────────────────────────────────────────────────────

    def _run(
        self,
        raw_path: str,
        label_map: Dict[str, str],
        output_dir: str,
        top_n: int,
        file_name: str,
    ) -> OmicsAnalysisResult:

        # 1. Parse Proteins sheet
        spc_df, int_df, _ = _parse_proteins_sheet(raw_path, label_map)

        if spc_df.empty and int_df.empty:
            raise ValueError("No SpC or Intensity columns found in 'Proteins' sheet.")

        # Prefer SpC for abundance; fall back to Intensity
        primary = spc_df if not spc_df.empty else int_df
        primary = primary.fillna(0)

        proteins_before = len(primary)

        # 2. Filter: remove rows that are all-zero
        primary = primary.loc[(primary > 0).any(axis=1)]
        proteins_after = len(primary)

        # 3. Log2 transform
        log2_mat = _log2_transform(primary)

        # 4. Build contrasts from whatever groups are present
        active_contrasts = _build_contrasts(label_map, set(log2_mat.columns))
        logger.info("Active contrasts: %s", [c[2] for c in active_contrasts])

        # 5. Fold changes
        fc_df = _compute_fold_changes(log2_mat, active_contrasts)

        # 6. Top 50 by variance
        top50_mat = _top_n_by_variance(log2_mat, n=min(top_n, len(log2_mat)))

        # 7. Rank proteins by activity score (sum of positive FCs across all contrasts)
        score = _rescue_score(fc_df)
        if score.sum() == 0 and not fc_df.empty:
            # All FCs are zero or negative — fall back to total absolute FC
            score = fc_df.abs().sum(axis=1)
        top_ranked = score.nlargest(top_n).index
        top_biomarkers: List[Dict[str, Any]] = []
        for rank, protein in enumerate(top_ranked, 1):
            entry: Dict[str, Any] = {"rank": rank, "protein": protein}
            if protein in fc_df.index:
                for col in fc_df.columns:
                    entry[col] = round(float(fc_df.at[protein, col]), 4)
            entry["rescue_score"] = round(float(score.get(protein, 0.0)), 4)
            top_biomarkers.append(entry)

        # 7. Save CSVs
        cleaned_path = str(Path(output_dir) / f"{file_name}_cleaned_expression.csv")
        fc_path      = str(Path(output_dir) / f"{file_name}_fold_changes.csv")
        top50_path   = str(Path(output_dir) / f"{file_name}_top50_rescued.csv")

        log2_mat.to_csv(cleaned_path)
        fc_df.to_csv(fc_path)

        top50_rescued = fc_df.loc[fc_df.index.isin(top_ranked)] if not fc_df.empty else fc_df
        top50_rescued.to_csv(top50_path)

        # 8. Plots
        plot_paths: List[str] = []

        heatmap_path = str(Path(output_dir) / f"{file_name}_heatmap.png")
        p = _plot_heatmap(top50_mat, heatmap_path)
        if p:
            plot_paths.append(p)

        for _, _, contrast_name in active_contrasts:
            if contrast_name in fc_df.columns:
                safe_contrast = re.sub(r"[^\w]", "_", str(contrast_name))
                bar_path = str(Path(output_dir) / f"{file_name}_barchart_{safe_contrast}.png")
                p = _plot_barchart(fc_df[contrast_name], contrast_name, bar_path)
                if p:
                    plot_paths.append(p)

        # Scatter: first two groups alphabetically (skip if fewer than 2 groups)
        groups_present = sorted(primary.columns.tolist())
        if len(groups_present) >= 2:
            g_first, g_second = groups_present[0], groups_present[1]
            safe_g1 = re.sub(r"[^\w]", "_", str(g_first))
            safe_g2 = re.sub(r"[^\w]", "_", str(g_second))
            scatter_path = str(Path(output_dir) / f"{file_name}_scatter_{safe_g1}_vs_{safe_g2}.png")
            p = _plot_scatter(primary, g_first, g_second, scatter_path)
            if p:
                plot_paths.append(p)

        venn_path = str(Path(output_dir) / f"{file_name}_venn_rescued.png")
        p = _plot_venn(fc_df, venn_path)
        if p:
            plot_paths.append(p)

        # 9. Excel report
        excel_path = str(Path(output_dir) / f"{file_name}_biomarker_report.xlsx")
        _write_excel_report(log2_mat, fc_df, top50_mat, top_biomarkers, excel_path)

        # 10. Count "significant" = |log2FC| >= threshold across any contrast
        n_sig = 0
        if not fc_df.empty:
            n_sig = int((fc_df.abs() >= _LOG2FC_THRESHOLD).any(axis=1).sum())

        qc_summary = {
            "proteins_before_filter": proteins_before,
            "proteins_after_qc":      proteins_after,
            "proteins_removed":       proteins_before - proteins_after,
            "groups_detected":        list(primary.columns),
            "contrasts_computed":     list(fc_df.columns),
            "log2_transformed":       True,
            "plot_paths":             plot_paths,
            "csv_paths":              [cleaned_path, fc_path, top50_path],
            # Pooled design note — no replicate-based QC applies
            "qc_note": (
                "Data is pre-normalised (MS LFQ / Olink NPX). "
                "Standard replicate-based QC is not applicable for pooled n=1 designs. "
                f"{proteins_before - proteins_after} contaminant / all-zero rows removed. "
                "Log₂ transformation applied with pseudocount +1."
            ),
            "qc_type": "pooled_prefilter",
        }

        code = self._generate_code(raw_path, label_map, active_contrasts, output_dir, top_n, file_name)

        logger.info(
            "PooledFoldChangeSkill done: %d proteins, %d contrasts, %d plots",
            proteins_after, len(fc_df.columns), len(plot_paths),
        )

        return OmicsAnalysisResult(
            omic_type=self.omic_type,
            top_biomarkers=top_biomarkers,
            n_significant=n_sig,
            excel_path=excel_path,
            qc_summary={**qc_summary, "analysis_code": code},
            error=None,
            analysis_code=code,
        )

    # ── Code generation ───────────────────────────────────────────────────────

    @staticmethod
    def _generate_code(
        raw_path: str,
        label_map: Dict[str, str],
        active_contrasts: List[Tuple[str, str, str]],
        output_dir: str,
        top_n: int,
        file_name: str,
    ) -> str:
        """Return a self-contained, re-executable Python script for this analysis."""
        L: List[str] = []
        a = L.append

        contrasts_repr = repr(active_contrasts)

        a('#!/usr/bin/env python3')
        a('"""')
        a('Reproducible pooled fold-change proteomics analysis')
        a('Auto-generated — edit parameters and re-run to reproduce results.')
        a('"""')
        a('')
        a('import numpy as np')
        a('import pandas as pd')
        a('import matplotlib; matplotlib.use("Agg")')
        a('import matplotlib.pyplot as plt')
        a('from pathlib import Path')
        a('')
        a('# ── Parameters (edit to customise) ──────────────────────────────────')
        a('RAW_PATH   = ' + repr(raw_path))
        a('LABEL_MAP  = ' + repr(label_map))
        a('OUTPUT_DIR = ' + repr(output_dir))
        a('TOP_N      = ' + str(top_n))
        a('FILE_NAME  = ' + repr(file_name))
        a('PSEUDOCOUNT = 1.0')
        a('LOG2FC_THRESHOLD = 1.0')
        a('')
        a('# Contrasts: (numerator_label, denominator_label, contrast_name)')
        a('# These are the exact contrasts computed from the uploaded data.')
        a('CONTRASTS = ' + contrasts_repr)
        a('')
        a('# ── 1. Parse Proteins sheet ──────────────────────────────────────────')
        a('xl = pd.ExcelFile(RAW_PATH, engine="openpyxl")')
        a('sheet = "Proteins" if "Proteins" in xl.sheet_names else xl.sheet_names[0]')
        a('df = xl.parse(sheet)')
        a('_ID_KEYWORDS = ("majority protein", "identified protein", "accession",')
        a('                "uniprot", "gene name", "protein id", "protein name")')
        a('id_col = next((c for c in df.columns')
        a('               if any(kw in str(c).lower() for kw in _ID_KEYWORDS)),')
        a('              df.columns[0])')
        a('df = df.set_index(id_col)')
        a('df.index = df.index.astype(str).str.strip()')
        a('')
        a('# Remove contaminants')
        a('mask = ~(df.index.str.startswith("REV__") | df.index.str.startswith("CON__"))')
        a('df = df[mask]')
        a('')
        a('# Build SpC matrix from label map')
        a('spc_cols = {}')
        a('for col in df.columns:')
        a('    cl = str(col).lower()')
        a('    for code, grp in LABEL_MAP.items():')
        a('        prefix = code.lower()')
        a('        is_match = (cl.startswith(f"{prefix} ") or cl.startswith(f"{prefix}_")')
        a('                    or f" {prefix} " in cl)')
        a('        if is_match and ("spectral" in cl or "spc" in cl or "ms/ms" in cl):')
        a('            spc_cols.setdefault(grp, col)')
        a('primary = pd.DataFrame({g: pd.to_numeric(df[c], errors="coerce").fillna(0)')
        a('                        for g, c in spc_cols.items()}, index=df.index)')
        a('primary = primary.loc[(primary > 0).any(axis=1)]')
        a('print(f"Proteins after filter: {len(primary)}")')
        a('')
        a('# ── 2. Log2 transform ────────────────────────────────────────────────')
        a('log2_mat = np.log2(primary + PSEUDOCOUNT)')
        a('')
        a('# ── 3. Fold changes ──────────────────────────────────────────────────')
        a('fc_records = {}')
        a('for num, den, name in CONTRASTS:')
        a('    if num in log2_mat.columns and den in log2_mat.columns:')
        a('        fc_records[name] = log2_mat[num] - log2_mat[den]')
        a('fc_df = pd.DataFrame(fc_records, index=log2_mat.index)')
        a('')
        a('# ── 4. Activity score: sum of positive FCs across all contrasts ──────')
        a('score = pd.Series(0.0, index=fc_df.index)')
        a('for col in fc_df.columns:')
        a('    score += fc_df[col].clip(lower=0)')
        a('if score.sum() == 0:  # fallback: rank by total absolute FC')
        a('    score = fc_df.abs().sum(axis=1)')
        a('top_proteins = score.nlargest(TOP_N).index')
        a('')
        a('# ── 5. Save outputs ──────────────────────────────────────────────────')
        a('Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)')
        a('log2_mat.to_csv(str(Path(OUTPUT_DIR) / f"{FILE_NAME}_cleaned_expression.csv"))')
        a('fc_df.to_csv(str(Path(OUTPUT_DIR) / f"{FILE_NAME}_fold_changes.csv"))')
        a('fc_df.loc[fc_df.index.isin(top_proteins)].to_csv(')
        a('    str(Path(OUTPUT_DIR) / f"{FILE_NAME}_top50_rescued.csv"))')
        a('print("Saved: cleaned expression, fold changes, top rescued proteins")')
        a('print(fc_df.loc[top_proteins].head(10).to_string())')

        return '\n'.join(L)
